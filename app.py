"""
Global Technical Analysis Dashboard  v5
=========================================
New in v5:
  - ATR (14-day daily) in screener tables and asset detail page
  - Signal history stored to GitHub (JSON file in repo); change detection
  - Page 4: Morning Bulletin — highlights, signal changes, strong signals, divergences
  - PDF export of bulletin (ReportLab, 1-2 pages)
  - Excel export of full screener table (openpyxl)
  - Volume bar subplot added to Page 3 charts

Setup for signal history (one-time):
  1. GitHub: Settings → Developer Settings → Personal Access Tokens → Fine-grained
     Permissions: Contents = Read & Write on your dashboard repo
  2. Streamlit Cloud: App settings → Secrets:
       GITHUB_TOKEN = "ghp_your_token_here"
       GITHUB_REPO  = "yourusername/your-repo-name"
       GITHUB_PATH  = "data/signal_history.json"   # will be created automatically

Run: streamlit run app.py
"""

import io, time, json, base64, urllib.request
from scipy.signal import find_peaks
import streamlit as st
import pandas as pd
import pandas_ta as ta
import yfinance as yf
import plotly.graph_objects as go
from plotly.subplots import make_subplots
from datetime import datetime, timedelta

# ReportLab imports
from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer, Table,
                                 TableStyle, HRFlowable, KeepTogether)
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import cm
from reportlab.lib import colors as rl_colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT

st.set_page_config(
    page_title="Global TA Dashboard",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ─────────────────────────────────────────────────────────────────────────────
# ASSET REGISTRIES
# ─────────────────────────────────────────────────────────────────────────────

INDICES = {
    "BIST 100":       {"ticker": "XU100.IS",  "currency": "₺", "region": "🇹🇷 Turkey"},
    "BIST 30":        {"ticker": "XU030.IS",  "currency": "₺", "region": "🇹🇷 Turkey"},
    "S&P 500":        {"ticker": "^GSPC",     "currency": "$", "region": "🇺🇸 US"},
    "NASDAQ 100":     {"ticker": "^NDX",      "currency": "$", "region": "🇺🇸 US"},
    "Dow Jones":      {"ticker": "^DJI",      "currency": "$", "region": "🇺🇸 US"},
    "EuroStoxx 50":   {"ticker": "^STOXX50E", "currency": "€", "region": "🇪🇺 Europe"},
    "DAX":            {"ticker": "^GDAXI",    "currency": "€", "region": "🇩🇪 Germany"},
    "CAC 40":         {"ticker": "^FCHI",     "currency": "€", "region": "🇫🇷 France"},
    "FTSE 100":       {"ticker": "^FTSE",     "currency": "£", "region": "🇬🇧 UK"},
    "Nikkei 225":     {"ticker": "^N225",     "currency": "¥", "region": "🇯🇵 Japan"},
    "Hang Seng":      {"ticker": "^HSI",      "currency": "HK$","region": "🇭🇰 HK"},
    "Shanghai Comp":  {"ticker": "000001.SS", "currency": "¥", "region": "🇨🇳 China"},
    "KOSPI":          {"ticker": "^KS11",     "currency": "₩", "region": "🇰🇷 Korea"},
    "ASX 200":        {"ticker": "^AXJO",     "currency": "A$","region": "🇦🇺 Australia"},
    "SENSEX":         {"ticker": "^BSESN",    "currency": "₹", "region": "🇮🇳 India"},
}
COMMODITIES = {
    "Gold":        {"ticker": "GC=F",  "region": "⛏️ Metals"},
    "Silver":      {"ticker": "SI=F",  "region": "⛏️ Metals"},
    "Copper":      {"ticker": "HG=F",  "region": "⛏️ Metals"},
    "Platinum":    {"ticker": "PL=F",  "region": "⛏️ Metals"},
    "Oil WTI":     {"ticker": "CL=F",  "region": "🛢️ Energy"},
    "Oil Brent":   {"ticker": "BZ=F",  "region": "🛢️ Energy"},
    "Natural Gas": {"ticker": "NG=F",  "region": "🛢️ Energy"},
    "Wheat":       {"ticker": "ZW=F",  "region": "🌾 Agri"},
    "Corn":        {"ticker": "ZC=F",  "region": "🌾 Agri"},
    "Soybeans":    {"ticker": "ZS=F",  "region": "🌾 Agri"},
    "Coffee":      {"ticker": "KC=F",  "region": "🌾 Agri"},
    "Sugar":       {"ticker": "SB=F",  "region": "🌾 Agri"},
}
FOREX = {
    "EUR/USD":         {"ticker": "EURUSD=X", "region": "💱 Forex"},
    "GBP/USD":         {"ticker": "GBPUSD=X", "region": "💱 Forex"},
    "USD/JPY":         {"ticker": "JPY=X",    "region": "💱 Forex"},
    "USD/TRY":         {"ticker": "TRY=X",    "region": "💱 Forex"},
    "USD/CNY":         {"ticker": "CNY=X",    "region": "💱 Forex"},
    "USD/CHF":         {"ticker": "CHF=X",    "region": "💱 Forex"},
    "USD Index (DXY)": {"ticker": "DX-Y.NYB", "region": "💱 Forex"},
}
CRYPTO = {
    "Bitcoin":  {"ticker": "BTC-USD", "region": "₿ Crypto"},
    "Ethereum": {"ticker": "ETH-USD", "region": "₿ Crypto"},
    "BNB":      {"ticker": "BNB-USD", "region": "₿ Crypto"},
    "Solana":   {"ticker": "SOL-USD", "region": "₿ Crypto"},
    "XRP":      {"ticker": "XRP-USD", "region": "₿ Crypto"},
}

BIST100_FALLBACK = {
    "AEFES.IS":"Anadolu Efes",    "AKBNK.IS":"Akbank",
    "AKSA.IS":"Aksa Akrilik",     "AKSEN.IS":"Aksa Enerji",
    "ALARK.IS":"Alarko Holding",  "ANSGR.IS":"Anadolu Sigorta",
    "ARCLK.IS":"Arçelik",         "ASELS.IS":"Aselsan",
    "ASTOR.IS":"Astor Enerji",    "BIMAS.IS":"BIM Magazalar",
    "BRSAN.IS":"Borusan Birlesik","BRYAT.IS":"Borusan Yatirim",
    "BSOKE.IS":"Batisoke",        "BTCIM.IS":"Baticim",
    "CCOLA.IS":"Coca Cola Icecek","CIMSA.IS":"Cimsa",
    "CVKMD.IS":"CVK Maden",       "CWENE.IS":"CW Enerji",
    "DAPGM.IS":"DAP GYO",         "DOAS.IS":"Dogus Otomotiv",
    "DOHOL.IS":"Dogan Holding",   "ECILC.IS":"Eczacibasi Ilac",
    "EFOR.IS":"Efor Yatirim",     "EKGYO.IS":"Emlak Konut GYO",
    "ENERYA.IS":"Enerya Enerji",  "ENJSA.IS":"Enerjisa Enerji",
    "ENKAI.IS":"Enka Insaat",     "EREGL.IS":"Eregil Demir Celik",
    "EUPWR.IS":"Europower Enerji","FENER.IS":"Fenerbahce",
    "FROTO.IS":"Ford Otosan",     "GARAN.IS":"Garanti BBVA",
    "GENIL.IS":"Gen Ilac",        "GLRMK.IS":"Gulermak",
    "GRSEL.IS":"Gursel Turizm",   "GSRAY.IS":"Galatasaray",
    "GUBRF.IS":"Gubre Fabrikalari","GWIND.IS":"Galata Wind",
    "HALKB.IS":"Halkbank",        "HEKTS.IS":"Hektas",
    "IPEKE.IS":"Ipek Dogal Enerji","ISCTR.IS":"Is Bankasi C",
    "ISGSY.IS":"Is Yatirim",      "IZMDC.IS":"Izdemir Enerji",
    "KATEV.IS":"Katilimevim",     "KCHOL.IS":"Koc Holding",
    "KMPUR.IS":"Kontrolmatik",    "KOZAA.IS":"Koza Anadolu Metal",
    "KOZAL.IS":"Turk Altin Isletmeleri","KRDMD.IS":"Kardemir D",
    "KUYAG.IS":"Kuyas Yatirim",   "MAGEN.IS":"Margun Enerji",
    "MAVI.IS":"Mavi Giyim",       "MGROS.IS":"Migros",
    "MIATK.IS":"Mia Teknoloji",   "MPARK.IS":"MLP Saglik",
    "ODAS.IS":"ODAS Elektrik",    "OTKAR.IS":"Otokar",
    "PAGYO.IS":"Pasifik GYO",     "PETKM.IS":"Petkim",
    "PGSUS.IS":"Pegasus",         "QUAGR.IS":"Qua Granite",
    "REEDR.IS":"Reeder Teknoloji","SAHOL.IS":"Sabanci Holding",
    "SARKY.IS":"Sarkuysan",       "SASA.IS":"SASA Polyester",
    "SISE.IS":"Sisecam",          "SKBNK.IS":"Sekerbank",
    "SOKM.IS":"Sok Marketler",    "TABGD.IS":"Tab Gida",
    "TAVHL.IS":"TAV Havalimanlari","TCELL.IS":"Turkcell",
    "THYAO.IS":"Turk Hava Yollari","TKFEN.IS":"Tekfen Holding",
    "TOASO.IS":"Tofas Oto",       "TSKB.IS":"TSKB",
    "TTKOM.IS":"Turk Telekom",    "TUKAS.IS":"Tukas Gida",
    "TUPRS.IS":"Tupras",          "ULKER.IS":"Ulker Biskuvi",
    "VAKBN.IS":"Vakifbank",       "VESTL.IS":"Vestel",
    "YKBNK.IS":"Yapi Kredi",      "ZOREN.IS":"Zorlu Enerji",
    "ALTIN.IS":"Altinay Savunma", "BALSU.IS":"Balsu Gida",
    "DSFKTR.IS":"Destek Finans",  "GRTHO.IS":"Grainturk Holding",
    "KLGYO.IS":"Kiler Holding",   "OBAMD.IS":"Oba Makarna",
    "PASEU.IS":"Pasifik Eurasia", "PSKGYO.IS":"Pasifik GYO",
    "TURKS.IS":"Tureks Turizm",   "YAZIC.IS":"AG Anadolu Group",
    "MRDIN.IS":"Oyak Cimento",    "GUSGR.IS":"Turkiye Sigorta",
    "CAN2T.IS":"Can2 Termik",
}
SP500_FALLBACK = {
    "AAPL":"Apple","MSFT":"Microsoft","NVDA":"NVIDIA","AMZN":"Amazon",
    "META":"Meta","GOOGL":"Alphabet A","GOOG":"Alphabet C","BRK-B":"Berkshire B",
    "LLY":"Eli Lilly","JPM":"JPMorgan","V":"Visa","XOM":"ExxonMobil",
    "UNH":"UnitedHealth","TSLA":"Tesla","MA":"Mastercard","PG":"P&G",
    "COST":"Costco","JNJ":"J&J","HD":"Home Depot","MRK":"Merck",
    "ABBV":"AbbVie","CVX":"Chevron","CRM":"Salesforce","BAC":"Bank of America",
    "WMT":"Walmart","NFLX":"Netflix","AMD":"AMD","KO":"Coca-Cola",
    "PEP":"PepsiCo","TMO":"Thermo Fisher","ACN":"Accenture","MCD":"McDonald's",
    "ORCL":"Oracle","CSCO":"Cisco","ABT":"Abbott","LIN":"Linde",
    "ADBE":"Adobe","GE":"GE Aerospace","DHR":"Danaher","TXN":"Texas Instruments",
    "PM":"Philip Morris","AMGN":"Amgen","CAT":"Caterpillar","INTU":"Intuit",
    "IBM":"IBM","SPGI":"S&P Global","NOW":"ServiceNow","GS":"Goldman Sachs",
    "ISRG":"Intuitive Surgical","BKNG":"Booking Holdings",
}

FREQ_OPTIONS = {
    "1 Hour":  {"period": "60d",  "interval": "1h"},
    "1 Day":   {"period": "2y",   "interval": "1d"},
    "1 Week":  {"period": "5y",   "interval": "1wk"},
}
SIG_COLORS = {
    "Strong Buy": "#00C853", "Buy": "#26A69A",
    "Hold": "#FFD740", "Sell": "#FF6D00",
    "Strong Sell": "#D50000", "N/A": "#555555",
}
DIV_LABELS = {
    "bullish":        ("Bull",      "#00C853"),
    "bearish":        ("Bear",      "#D50000"),
    "hidden_bullish": ("H.Bull",    "#26A69A"),
    "hidden_bearish": ("H.Bear",    "#FF6D00"),
    None:             ("—",         "#555555"),
}
STRENGTH_LABEL = {0:"", 1:"●", 2:"●●", 3:"●●●"}

# ─────────────────────────────────────────────────────────────────────────────
# CSS
# ─────────────────────────────────────────────────────────────────────────────
CSS = """
<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700&display=swap');
html,body,[class*="css"]{font-family:'Inter',sans-serif;}
.hdr{background:linear-gradient(135deg,#1a1f3a,#0d1117);border-bottom:2px solid #2E75B6;
  padding:.9rem 2rem .7rem;margin:-1rem -1rem 1rem -1rem;}
.hdr h1{margin:0;font-size:1.4rem;font-weight:700;
  background:linear-gradient(90deg,#63b3ed,#68d391);
  -webkit-background-clip:text;-webkit-text-fill-color:transparent;}
.hdr p{margin:.1rem 0 0;color:#718096;font-size:.76rem;}
.ss{font-size:.66rem;font-weight:700;letter-spacing:.1em;color:#63b3ed;
  text-transform:uppercase;padding:.5rem 0 .1rem;border-top:1px solid #2d3748;margin-top:.4rem;}
.mc{background:#161b2e;border:1px solid #2d3748;border-radius:8px;
  padding:.7rem .85rem;text-align:center;}
.mc .lb{font-size:.62rem;color:#718096;text-transform:uppercase;
  letter-spacing:.07em;margin-bottom:.18rem;}
.mc .vl{font-size:1.15rem;font-weight:700;}
.mc .sb{font-size:.65rem;color:#718096;margin-top:.1rem;}
.sig{display:inline-block;padding:.18rem .6rem;border-radius:20px;
  font-size:.72rem;font-weight:600;letter-spacing:.03em;}
.sig-strong-buy {background:#003d1f;color:#00C853;border:1px solid #00C853;}
.sig-buy        {background:#0d2e2a;color:#26A69A;border:1px solid #26A69A;}
.sig-hold       {background:#3a3000;color:#FFD740;border:1px solid #FFD740;}
.sig-sell       {background:#3a1a00;color:#FF6D00;border:1px solid #FF6D00;}
.sig-strong-sell{background:#3a0000;color:#FF5252;border:1px solid #FF5252;}
.sig-na         {background:#1a1a1a;color:#888;border:1px solid #444;}
.srcnote{font-size:.68rem;color:#4a5568;border:1px solid #2d3748;
  border-radius:5px;padding:.35rem .7rem;margin-bottom:.5rem;}
/* Bulletin cards */
.bcard{background:#0d1117;border:1px solid #2d3748;border-radius:10px;
  padding:.8rem 1rem;margin-bottom:.6rem;}
.bcard .btype{font-size:.64rem;color:#718096;text-transform:uppercase;
  letter-spacing:.08em;margin-bottom:.2rem;}
.bcard .btitle{font-size:.95rem;font-weight:600;margin-bottom:.15rem;}
.bcard .bsub{font-size:.75rem;color:#a0aec0;}
/* Change arrow */
.chg-up{color:#00C853;font-weight:700;}
.chg-dn{color:#D50000;font-weight:700;}
.footer{text-align:center;color:#4a5568;font-size:.68rem;
  padding:.9rem 0;border-top:1px solid #1a202c;margin-top:1rem;}
</style>
"""

# ─────────────────────────────────────────────────────────────────────────────
# GITHUB SIGNAL HISTORY
# ─────────────────────────────────────────────────────────────────────────────

def _gh_request(method: str, path: str, data: dict = None) -> dict:
    """Make a GitHub API request using secrets from Streamlit."""
    try:
        token = st.secrets.get("GITHUB_TOKEN", "")
        repo  = st.secrets.get("GITHUB_REPO",  "")
        if not token or not repo:
            return {"error": "no_secrets"}
        url = f"https://api.github.com/repos/{repo}/contents/{path}"
        hdrs = {
            "Authorization": f"token {token}",
            "Accept": "application/vnd.github.v3+json",
            "Content-Type": "application/json",
        }
        body = json.dumps(data).encode() if data else None
        req  = urllib.request.Request(url, data=body, headers=hdrs, method=method)
        with urllib.request.urlopen(req, timeout=10) as r:
            return json.loads(r.read())
    except Exception as e:
        return {"error": str(e)}


def load_signal_history() -> dict:
    """
    Load signal history JSON from GitHub.
    Format: { "YYYY-MM-DD": { "TICKER": {"signal":"..","score":N,"atr_pct":X} } }
    Returns empty dict if not configured or file doesn't exist yet.
    """
    gh_path = st.secrets.get("GITHUB_PATH", "data/signal_history.json")
    resp = _gh_request("GET", gh_path)
    if "error" in resp or "content" not in resp:
        return {}
    try:
        content = base64.b64decode(resp["content"]).decode("utf-8")
        return json.loads(content)
    except Exception:
        return {}


def save_signal_history(history: dict) -> bool:
    """Save updated signal history JSON to GitHub (creates or updates file)."""
    gh_path = st.secrets.get("GITHUB_PATH", "data/signal_history.json")
    content_b64 = base64.b64encode(
        json.dumps(history, indent=2).encode("utf-8")
    ).decode("utf-8")

    # Need SHA to update existing file
    existing = _gh_request("GET", gh_path)
    sha = existing.get("sha")

    data = {
        "message": f"Signal history update {datetime.now().strftime('%Y-%m-%d %H:%M')}",
        "content": content_b64,
    }
    if sha:
        data["sha"] = sha

    resp = _gh_request("PUT", gh_path, data)
    return "error" not in resp


def get_signal_changes(history: dict, today_signals: dict) -> list:
    """
    Compare today's signals with the most recent previous date in history.
    Returns list of changes sorted by magnitude (biggest change first).
    """
    if not history:
        return []
    dates = sorted(history.keys())
    # Find the most recent date that is NOT today
    today_str = datetime.now().strftime("%Y-%m-%d")
    prev_dates = [d for d in dates if d < today_str]
    if not prev_dates:
        return []
    prev_day = history[prev_dates[-1]]

    sig_order = {"Strong Buy":5,"Buy":4,"Hold":3,"Sell":2,"Strong Sell":1,"N/A":0}
    changes = []
    for ticker, td in today_signals.items():
        prev = prev_day.get(ticker)
        if not prev:
            continue
        ps, cs = prev.get("signal","N/A"), td.get("signal","N/A")
        if ps == cs:
            continue
        direction = "up" if sig_order.get(cs,0) > sig_order.get(ps,0) else "down"
        changes.append({
            "ticker":     ticker,
            "name":       td.get("name",""),
            "prev_sig":   ps,
            "curr_sig":   cs,
            "prev_score": prev.get("score",0),
            "curr_score": td.get("score",0),
            "direction":  direction,
        })
    return sorted(changes, key=lambda x: abs(x["curr_score"]-x["prev_score"]), reverse=True)

# ─────────────────────────────────────────────────────────────────────────────
# CONSTITUENT FETCHING
# ─────────────────────────────────────────────────────────────────────────────

@st.cache_data(ttl=86400, show_spinner=False)
def fetch_bist100_live() -> dict:
    import re, ssl
    try:
        url  = "https://www.borsaistanbul.com/datum/hisse_endeks_ds.csv"
        hdrs = {
            "User-Agent": ("Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                           "AppleWebKit/537.36 Chrome/124.0.0.0 Safari/537.36"),
            "Accept": "text/csv,*/*",
            "Referer": "https://www.borsaistanbul.com/en/indices/bist-stock-indices",
        }
        ctx = ssl.create_default_context()
        ctx.check_hostname = False; ctx.verify_mode = ssl.CERT_NONE
        req = urllib.request.Request(url, headers=hdrs)
        with urllib.request.urlopen(req, timeout=12, context=ctx) as r:
            raw = r.read()
        for enc in ("utf-8-sig","utf-8","windows-1252","iso-8859-9"):
            try: text = raw.decode(enc); break
            except: pass
        else: return BIST100_FALLBACK
        df = pd.read_csv(io.StringIO(text), sep=";", dtype=str, on_bad_lines="skip")
        df.columns = [c.strip().upper() for c in df.columns]
        ticker_col = next((c for c in df.columns if any(k in c for k in ("TICKER","KOD","SEMBOL","SYMBOL","CODE"))), None)
        index_col  = next((c for c in df.columns if any(k in c for k in ("ENDEKS","INDEX","INDICE"))), None)
        name_col   = next((c for c in df.columns if any(k in c for k in ("ISIM","NAME","SIRKET","UNVAN","COMPANY","ACIKLAMA"))), None)
        if not ticker_col or not index_col: return BIST100_FALLBACK
        bdf = df[df[index_col].str.strip().str.upper() == "XU100"]
        if bdf.empty: return BIST100_FALLBACK
        result = {}
        for _, row in bdf.iterrows():
            raw_t = str(row[ticker_col]).strip().upper()
            clean = re.sub(r'\.[A-Z]{1,2}$', '', raw_t)
            name  = str(row[name_col]).strip() if name_col and name_col in row else clean
            result[f"{clean}.IS"] = name
        return result if len(result) >= 50 else BIST100_FALLBACK
    except Exception:
        return BIST100_FALLBACK


@st.cache_data(ttl=86400, show_spinner=False)
def fetch_sp500_live() -> dict:
    try:
        url = ("https://raw.githubusercontent.com/datasets/"
               "s-and-p-500-companies/main/data/constituents.csv")
        req = urllib.request.Request(url, headers={"User-Agent":"Mozilla/5.0"})
        with urllib.request.urlopen(req, timeout=10) as r:
            text = r.read().decode("utf-8")
        df   = pd.read_csv(io.StringIO(text))
        sym  = "Symbol"   if "Symbol"   in df.columns else df.columns[0]
        name = "Security" if "Security" in df.columns else df.columns[1]
        result = {str(r[sym]).replace(".","–"): str(r[name]) for _,r in df.iterrows()}
        return result if len(result) >= 400 else SP500_FALLBACK
    except Exception:
        return SP500_FALLBACK

# ─────────────────────────────────────────────────────────────────────────────
# PRICE DATA
# ─────────────────────────────────────────────────────────────────────────────

@st.cache_data(ttl=3600, show_spinner=False)
def fetch_data(full_ticker: str, period: str, interval: str) -> pd.DataFrame:
    for _ in range(2):
        try:
            df = yf.download(full_ticker, period=period, interval=interval,
                             auto_adjust=True, progress=False, threads=False)
            if df is None or df.empty: time.sleep(0.5); continue
            if isinstance(df.columns, pd.MultiIndex):
                lvl0 = df.columns.get_level_values(0).astype(str)
                lvl1 = df.columns.get_level_values(1).astype(str)
                fields = {"Open","High","Low","Close","Volume","Adj Close"}
                df.columns = lvl0 if len(fields & set(lvl0.values)) >= 3 else lvl1
            df.columns = [str(c).lower() for c in df.columns]
            if not {"open","high","low","close"}.issubset(set(df.columns)):
                time.sleep(0.5); continue
            df.index = pd.to_datetime(df.index)
            df = df.dropna(subset=["close"])
            if not df.empty: return df
            time.sleep(0.5)
        except Exception: time.sleep(0.5)
    return pd.DataFrame()


@st.cache_data(ttl=86400, show_spinner=False)
def fetch_atr_daily(full_ticker: str) -> dict:
    """
    Always fetch 14-day ATR on DAILY timeframe regardless of selected timeframe.
    Returns {'atr': float, 'atr_pct': float} or empty dict.
    """
    try:
        df = yf.download(full_ticker, period="3mo", interval="1d",
                         auto_adjust=True, progress=False, threads=False)
        if df is None or df.empty: return {}
        if isinstance(df.columns, pd.MultiIndex):
            lvl0 = df.columns.get_level_values(0).astype(str)
            fields = {"Open","High","Low","Close","Volume"}
            df.columns = lvl0 if len(fields & set(lvl0.values)) >= 3 else df.columns.get_level_values(1).astype(str)
        df.columns = [str(c).lower() for c in df.columns]
        if not {"high","low","close"}.issubset(set(df.columns)): return {}
        df = df.dropna(subset=["close"])
        atr_series = ta.atr(df["high"], df["low"], df["close"], length=14)
        if atr_series is None or atr_series.dropna().empty: return {}
        atr_val = float(atr_series.iloc[-1])
        price   = float(df["close"].iloc[-1])
        return {"atr": round(atr_val, 4), "atr_pct": round((atr_val / price) * 100, 2)}
    except Exception:
        return {}

# ─────────────────────────────────────────────────────────────────────────────
# DIVERGENCE DETECTION
# ─────────────────────────────────────────────────────────────────────────────

def _find_swings(series: pd.Series, order: int):
    s = series.dropna(); arr = s.values
    if len(arr) < order * 2 + 1:
        return pd.Series(dtype=float), pd.Series(dtype=float)
    hi, _ = find_peaks( arr, distance=order, prominence=0)
    lo, _ = find_peaks(-arr, distance=order, prominence=0)
    return (pd.Series(arr[hi], index=s.index[hi]),
            pd.Series(arr[lo], index=s.index[lo]))


def _bar_gap(series: pd.Series, idx1, idx2) -> int:
    try: return abs(series.index.get_loc(idx2) - series.index.get_loc(idx1))
    except: return 0


def _strength(diff: float, thr: float) -> int:
    if diff >= thr * 4: return 3
    if diff >= thr * 2: return 2
    return 1


def _check_pair(p_sw, i_sw, price, ind, min_bars, min_diff, bullish_direction):
    if len(p_sw) < 2 or len(i_sw) < 2: return {}
    p1i,p2i = p_sw.index[-2], p_sw.index[-1]
    p1, p2  = float(p_sw.iloc[-2]), float(p_sw.iloc[-1])
    i1i,i2i = i_sw.index[-2], i_sw.index[-1]
    i1, i2  = float(i_sw.iloc[-2]), float(i_sw.iloc[-1])
    if _bar_gap(price, p1i, p2i) < min_bars: return {}
    if abs(i1 - i2) < min_diff: return {}
    r = {}
    if bullish_direction:
        if p2 < p1 and i2 > i1:
            r = {"type":"bullish","strength":_strength(i2-i1,min_diff),
                 "p_x":[p1i,p2i],"p_y":[p1,p2],"i_x":[i1i,i2i],"i_y":[i1,i2]}
        elif p2 > p1 and i2 < i1:
            r = {"type":"hidden_bullish","strength":_strength(i1-i2,min_diff),
                 "p_x":[p1i,p2i],"p_y":[p1,p2],"i_x":[i1i,i2i],"i_y":[i1,i2]}
    else:
        if p2 > p1 and i2 < i1:
            r = {"type":"bearish","strength":_strength(i1-i2,min_diff),
                 "p_x":[p1i,p2i],"p_y":[p1,p2],"i_x":[i1i,i2i],"i_y":[i1,i2]}
        elif p2 < p1 and i2 > i1:
            r = {"type":"hidden_bearish","strength":_strength(i2-i1,min_diff),
                 "p_x":[p1i,p2i],"p_y":[p1,p2],"i_x":[i1i,i2i],"i_y":[i1,i2]}
    return r


def detect_divergence(price, ind, lookback=150, order=5, min_bars=10, min_diff=3.0) -> dict:
    empty = {"type":None,"strength":0}
    try:
        p  = price.iloc[-lookback:].dropna()
        i  = ind.iloc[-lookback:].dropna()
        cm = p.index.intersection(i.index)
        if len(cm) < order * 4: return empty
        p = p[cm]; i = i[cm]
        ph,pl = _find_swings(p, order)
        ih,il = _find_swings(i, order)
        for fn,bdir in [(_check_pair,[ph,ih,p,i,min_bars,min_diff,False]),
                        (_check_pair,[pl,il,p,i,min_bars,min_diff,True])]:
            r = fn(*bdir)
            if r.get("type") in ("bearish","bullish"): return r
        for fn,bdir in [(_check_pair,[ph,ih,p,i,min_bars,min_diff,False]),
                        (_check_pair,[pl,il,p,i,min_bars,min_diff,True])]:
            r = fn(*bdir)
            if r.get("type"): return r
    except Exception: pass
    return empty


def compute_divergences(df, ind, interval) -> dict:
    order    = 3 if interval == "1h" else 5
    min_bars = 5 if interval == "1h" else 10
    lookback = 100 if interval == "1h" else 150
    rsi    = ind.get("rsi", pd.Series(dtype=float))
    macd_h = ind.get("macd", pd.DataFrame()).get(ind.get("macdh_col",""), pd.Series(dtype=float))
    min_m  = float(macd_h.abs().quantile(0.25)) if len(macd_h.dropna()) > 10 else 0.1
    return {
        "rsi":  detect_divergence(df["close"], rsi,    lookback, order, min_bars, 3.0),
        "macd": detect_divergence(df["close"], macd_h, lookback, order, min_bars, min_m),
    }

# ─────────────────────────────────────────────────────────────────────────────
# INDICATORS & SCORING
# ─────────────────────────────────────────────────────────────────────────────

def _s_rsi(v):
    if pd.isna(v): return 0
    if v>=70: return -2
    if v>=60: return -1
    if v<=30: return  2
    if v<=40: return  1
    return 0

def _s_macd(m,s,h):
    if any(pd.isna(x) for x in [m,s,h]): return 0
    return max(-2, min(2,(1 if m>s else -1)+(1 if h>0 else -1)))

def _s_cci(v):
    if pd.isna(v): return 0
    if v>=200: return 2; if v>=100: return 1
    if v<=-200: return -2; if v<=-100: return -1
    return 0

def _s_mom_pct(pct):
    if pd.isna(pct): return 0
    if pct>3: return 2; if pct>0: return 1
    if pct<-3: return -2
    return -1

def _s_stoch(k,d):
    if any(pd.isna(x) for x in [k,d]): return 0
    if k<20 and d<20: return 2; if k>80 and d>80: return -2
    if k<20: return 1; if k>80: return -1
    return 1 if k>d else -1

def _s_willr(v):
    if pd.isna(v): return 0
    if v<=-80: return 2; if v<=-50: return 1
    if v>=-20: return -2
    return -1

def _label(t):
    if t>=6: return "Strong Buy"; if t>=2: return "Buy"
    if t<=-6: return "Strong Sell"; if t<=-2: return "Sell"
    return "Hold"


def compute_signal(df, p, full_series=False, interval="1d") -> dict:
    empty = {"signal":"N/A","score":0,"details":{},"ind_series":{},"divergences":{}}
    if len(df) < 35: return empty
    try:
        d = df.copy() if full_series else df.tail(200).copy()
        mf,ms,mg = p["macd_fast"],p["macd_slow"],p["macd_sig"]
        rl,ml,cl = p["rsi_len"],p["mom_len"],p["cci_len"]
        sk,sd,wl = p["stoch_k"],p["stoch_d"],p["willr_len"]

        macd_df = ta.macd(d["close"], fast=mf, slow=ms, signal=mg)
        rsi     = ta.rsi(d["close"], length=rl)
        mom     = ta.mom(d["close"], length=ml)
        cci     = ta.cci(d["high"],d["low"],d["close"], length=cl)
        stoch   = ta.stoch(d["high"],d["low"],d["close"], k=sk,d=sd)
        willr   = ta.willr(d["high"],d["low"],d["close"], length=wl)

        close_n_ago    = d["close"].shift(ml)
        mom_pct_series = (mom / close_n_ago.replace(0,float("nan"))) * 100

        macd_col  = f"MACD_{mf}_{ms}_{mg}"
        macds_col = f"MACDs_{mf}_{ms}_{mg}"
        macdh_col = f"MACDh_{mf}_{ms}_{mg}"
        stk_col   = f"STOCHk_{sk}_{sd}_{sd}"
        std_col   = f"STOCHd_{sk}_{sd}_{sd}"

        sm  = _s_macd(macd_df[macd_col].iloc[-1],macd_df[macds_col].iloc[-1],macd_df[macdh_col].iloc[-1])
        sr  = _s_rsi(rsi.iloc[-1])
        smo = _s_mom_pct(float(mom_pct_series.iloc[-1]))
        sc  = _s_cci(cci.iloc[-1])
        ss  = _s_stoch(stoch[stk_col].iloc[-1],stoch[std_col].iloc[-1])
        sw  = _s_willr(willr.iloc[-1])
        total = sr + sm + sc + smo + ss + sw

        ind_series_dict = {
            "macd":macd_df,"rsi":rsi,"mom_pct":mom_pct_series,
            "cci":cci,"stoch":stoch,"willr":willr,
            "macd_col":macd_col,"macds_col":macds_col,
            "macdh_col":macdh_col,"stk_col":stk_col,"std_col":std_col,
        }
        divs = compute_divergences(d, ind_series_dict, interval) if full_series else {}

        return {
            "signal":_label(total),"score":total,
            "ind_series":ind_series_dict,"divergences":divs,
            "details":{
                "RSI":       (sr,  round(float(rsi.iloc[-1]),2)),
                "MACD":      (sm,  round(float(macd_df[macd_col].iloc[-1]),4)),
                "CCI":       (sc,  round(float(cci.iloc[-1]),2)),
                "Momentum%": (smo, round(float(mom_pct_series.iloc[-1]),2)),
                "Stochastic":(ss,  round(float(stoch[stk_col].iloc[-1]),2)),
                "Williams%R":(sw,  round(float(willr.iloc[-1]),2)),
            },
        }
    except Exception: return empty


def one_row(full_ticker, name, atype, region, period, interval, p) -> dict:
    base = {"Ticker":full_ticker,"Name":name,"Type":atype,"Region":region,
            "Price":None,"Change%":None,"Signal":"N/A","Score":0,
            "ATR":None,"ATR%":None,
            "RSI":0,"MACD":0,"CCI":0,"Momentum%":0,"Stochastic":0,"Williams%R":0,
            "RSI Div":None,"MACD Div":None}
    df = fetch_data(full_ticker, period, interval)
    if df.empty or len(df) < 35: return base

    sig   = compute_signal(df, p, interval=interval)
    price = float(df["close"].iloc[-1])
    chg   = float((df["close"].iloc[-1]/df["close"].iloc[-2]-1)*100) if len(df)>1 else 0.0

    # ATR — always daily
    atr_data = fetch_atr_daily(full_ticker)

    # Divergences for screener
    try:
        d_t = df.tail(200).copy()
        mf,ms,mg = p["macd_fast"],p["macd_slow"],p["macd_sig"]
        rsi_s  = ta.rsi(d_t["close"], length=p["rsi_len"])
        macd_s = ta.macd(d_t["close"], fast=mf, slow=ms, signal=mg)
        mh_s   = macd_s[f"MACDh_{mf}_{ms}_{mg}"]
        mm_d   = float(mh_s.abs().quantile(0.25)) if len(mh_s.dropna())>10 else 0.1
        ordr   = 3 if interval=="1h" else 5
        mnb    = 5 if interval=="1h" else 10
        lb     = 100 if interval=="1h" else 150
        rsi_div  = detect_divergence(d_t["close"], rsi_s, lb, ordr, mnb, 3.0)
        macd_div = detect_divergence(d_t["close"], mh_s,  lb, ordr, mnb, mm_d)
    except Exception:
        rsi_div = macd_div = {"type":None,"strength":0}

    row = {**base,
           "Price":   round(price, 4 if price<10 else 2),
           "Change%": round(chg, 2),
           "Signal":  sig["signal"],
           "Score":   sig["score"],
           "ATR":     atr_data.get("atr"),
           "ATR%":    atr_data.get("atr_pct"),
           "RSI Div": rsi_div,
           "MACD Div":macd_div}
    for k,(sc,_) in sig.get("details",{}).items():
        row[k] = sc
    return row

# ─────────────────────────────────────────────────────────────────────────────
# CHARTS
# ─────────────────────────────────────────────────────────────────────────────

def build_charts(df, ind, ticker, name, interval, divs=None, atr_data=None) -> go.Figure:
    # 8 rows: price, volume, MACD, RSI, Mom%, CCI, Stoch, WillR
    fig = make_subplots(
        rows=8, cols=1, shared_xaxes=True, vertical_spacing=0.018,
        row_heights=[3, 0.8, 1.2, 1.2, 1.0, 1.0, 1.0, 1.0],
        subplot_titles=[
            f"{ticker} — {name}", "Volume",
            "MACD", "RSI (14)", "Momentum % (10)",
            "CCI (20)", "Stochastic (14/3)", "Williams %R (14)",
        ],
    )

    # ── Price (candlestick)
    fig.add_trace(go.Candlestick(
        x=df.index, open=df["open"], high=df["high"],
        low=df["low"], close=df["close"],
        increasing_line_color="#26A69A", decreasing_line_color="#EF5350",
        showlegend=False, name="Price"), row=1, col=1)

    # ATR bands on price chart (optional)
    if atr_data and atr_data.get("atr"):
        atr_val = atr_data["atr"]
        last_close = float(df["close"].iloc[-1])
        fig.add_hline(y=last_close + atr_val, line_dash="dot",
                      line_color="rgba(255,215,64,0.4)", line_width=1, row=1, col=1)
        fig.add_hline(y=last_close - atr_val, line_dash="dot",
                      line_color="rgba(255,215,64,0.4)", line_width=1, row=1, col=1)

    # ── Volume
    if "volume" in df.columns:
        vol_colors = ["#26A69A" if c >= o else "#EF5350"
                      for c, o in zip(df["close"], df["open"])]
        fig.add_trace(go.Bar(x=df.index, y=df["volume"],
            marker_color=vol_colors, showlegend=False, name="Volume"), row=2, col=1)
        # 20-bar avg volume line
        vol_ma = df["volume"].rolling(20).mean()
        fig.add_trace(go.Scatter(x=df.index, y=vol_ma,
            line=dict(color="#FFD740", width=1, dash="dot"),
            showlegend=False, name="Vol MA20"), row=2, col=1)

    # ── MACD
    md = ind["macd"]; hc,mc,sc_col = ind["macdh_col"],ind["macd_col"],ind["macds_col"]
    h  = md[hc]
    fig.add_trace(go.Bar(x=df.index, y=h, showlegend=False,
        marker_color=["#26A69A" if v>=0 else "#EF5350" for v in h]), row=3, col=1)
    fig.add_trace(go.Scatter(x=df.index, y=md[mc],
        line=dict(color="#2962FF",width=1.3), name="MACD"), row=3, col=1)
    fig.add_trace(go.Scatter(x=df.index, y=md[sc_col],
        line=dict(color="#FF6D00",width=1.2,dash="dot"), name="Signal"), row=3, col=1)

    # ── RSI
    fig.add_trace(go.Scatter(x=df.index, y=ind["rsi"],
        line=dict(color="#AB47BC",width=1.5), showlegend=False), row=4, col=1)
    for lv,c in [(70,"rgba(239,83,80,.4)"),(30,"rgba(38,166,154,.4)")]:
        fig.add_hline(y=lv, line_dash="dash", line_color=c, row=4, col=1)
    fig.add_hline(y=50, line_dash="dot", line_color="grey", line_width=.6, row=4, col=1)

    # ── Momentum %
    mp = ind["mom_pct"]
    fig.add_trace(go.Bar(x=df.index, y=mp, showlegend=False,
        marker_color=["#26A69A" if v>=0 else "#EF5350" for v in mp]), row=5, col=1)
    for lv,c in [(3,"rgba(239,83,80,.35)"),(-3,"rgba(38,166,154,.35)")]:
        fig.add_hline(y=lv, line_dash="dash", line_color=c, row=5, col=1)
    fig.add_hline(y=0, line_dash="dot", line_color="grey", line_width=.6, row=5, col=1)

    # ── CCI
    fig.add_trace(go.Scatter(x=df.index, y=ind["cci"],
        line=dict(color="#FFA726",width=1.5), showlegend=False), row=6, col=1)
    for lv,c in [(100,"rgba(239,83,80,.35)"),(-100,"rgba(38,166,154,.35)")]:
        fig.add_hline(y=lv, line_dash="dash", line_color=c, row=6, col=1)

    # ── Stochastic
    stk,std = ind["stk_col"],ind["std_col"]
    fig.add_trace(go.Scatter(x=df.index, y=ind["stoch"][stk],
        line=dict(color="#42A5F5",width=1.5), name="%K"), row=7, col=1)
    fig.add_trace(go.Scatter(x=df.index, y=ind["stoch"][std],
        line=dict(color="#EF5350",width=1.2,dash="dot"), name="%D"), row=7, col=1)
    for lv,c in [(80,"rgba(239,83,80,.35)"),(20,"rgba(38,166,154,.35)")]:
        fig.add_hline(y=lv, line_dash="dash", line_color=c, row=7, col=1)

    # ── Williams %R
    fig.add_trace(go.Scatter(x=df.index, y=ind["willr"],
        line=dict(color="#EC407A",width=1.5), showlegend=False), row=8, col=1)
    for lv,c in [(-20,"rgba(239,83,80,.35)"),(-80,"rgba(38,166,154,.35)")]:
        fig.add_hline(y=lv, line_dash="dash", line_color=c, row=8, col=1)

    # ── Divergence lines
    if divs:
        DIV_ROW = {"rsi":4,"macd":3}
        DIV_COLORS = {"bullish":"#00C853","hidden_bullish":"#26A69A",
                      "bearish":"#D50000","hidden_bearish":"#FF6D00"}
        for ind_key, row_num in DIV_ROW.items():
            d = divs.get(ind_key,{})
            dtype = d.get("type")
            if not dtype: continue
            color = DIV_COLORS.get(dtype,"#FFD740")
            px_arr,py_arr = d.get("p_x",[]),d.get("p_y",[])
            ix_arr,iy_arr = d.get("i_x",[]),d.get("i_y",[])
            if len(px_arr)<2 or len(ix_arr)<2: continue
            fig.add_shape(type="line",x0=str(px_arr[0]),y0=py_arr[0],
                x1=str(px_arr[1]),y1=py_arr[1],
                line=dict(color=color,width=2,dash="dot"),row=1,col=1)
            for xi,yi in zip(px_arr,py_arr):
                fig.add_scatter(x=[xi],y=[yi],mode="markers",
                    marker=dict(color=color,size=7),showlegend=False,row=1,col=1)
            fig.add_shape(type="line",x0=str(ix_arr[0]),y0=iy_arr[0],
                x1=str(ix_arr[1]),y1=iy_arr[1],
                line=dict(color=color,width=2,dash="dot"),row=row_num,col=1)
            for xi,yi in zip(ix_arr,iy_arr):
                fig.add_scatter(x=[xi],y=[yi],mode="markers",
                    marker=dict(color=color,size=7),showlegend=False,row=row_num,col=1)
            s = d.get("strength",1); dots = "●"*s
            lbl = {"bullish":"Bull","bearish":"Bear",
                   "hidden_bullish":"H.Bull","hidden_bearish":"H.Bear"}.get(dtype,"Div")
            fig.add_annotation(x=str(px_arr[1]),y=py_arr[1],
                text=f"<b>{lbl} {dots}</b>",
                font=dict(color=color,size=10),
                bgcolor="rgba(0,0,0,0.6)",bordercolor=color,
                borderwidth=1,borderpad=3,showarrow=False,yshift=14,
                row=1,col=1)

    # ── Layout
    rb_d = [dict(bounds=["sat","mon"])]
    rb_h = [dict(bounds=["sat","mon"]),dict(bounds=[21,9],pattern="hour")]
    rb   = rb_h if interval=="1h" else rb_d

    fig.update_layout(
        height=1280,
        paper_bgcolor="#0E1117", plot_bgcolor="#0E1117",
        font=dict(color="#FAFAFA",family="Inter,sans-serif",size=11),
        xaxis_rangeslider_visible=False,
        margin=dict(l=55,r=20,t=38,b=20),
        hovermode="x unified",
        legend=dict(orientation="h",x=0,y=1.01,bgcolor="rgba(0,0,0,0)",font=dict(size=10)),
    )
    for i in range(1,9):
        fig.update_xaxes(showgrid=True,gridcolor="#1E2130",gridwidth=.5,
                         rangebreaks=rb,row=i,col=1)
        fig.update_yaxes(showgrid=True,gridcolor="#1E2130",gridwidth=.5,
                         zeroline=False,row=i,col=1)
    return fig

# ─────────────────────────────────────────────────────────────────────────────
# PDF BULLETIN
# ─────────────────────────────────────────────────────────────────────────────

def _div_str_plain(cell) -> str:
    if not isinstance(cell, dict): return "—"
    dtype = cell.get("type")
    if not dtype: return "—"
    s = cell.get("strength",0)
    dots = {1:"●",2:"●●",3:"●●●"}.get(s,"")
    short = {"bullish":"▲ Bull","bearish":"▼ Bear",
             "hidden_bullish":"▲ H.Bull","hidden_bearish":"▼ H.Bear"}.get(dtype,"—")
    return f"{short} {dots}"


def build_bulletin_pdf(df_all: pd.DataFrame, freq: str,
                       changes: list, date_str: str) -> bytes:
    """
    Generate a 1-2 page morning bulletin PDF.
    Sections:
      1. Header (title, date, timeframe, signal summary)
      2. Signal changes since last run
      3. Strong Buy candidates
      4. Strong Sell candidates
      5. Active divergences
    """
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
        rightMargin=1.8*cm, leftMargin=1.8*cm,
        topMargin=1.8*cm, bottomMargin=1.8*cm)

    styles = getSampleStyleSheet()
    BLUE   = rl_colors.HexColor('#1F3864')
    LBLUE  = rl_colors.HexColor('#2E75B6')
    ACCENT = rl_colors.HexColor('#EEF4FB')
    GREEN  = rl_colors.HexColor('#00C853')
    RED    = rl_colors.HexColor('#D50000')
    AMBER  = rl_colors.HexColor('#FF6D00')
    GOLD   = rl_colors.HexColor('#FFD740')
    GRAY   = rl_colors.HexColor('#718096')

    H1 = ParagraphStyle('H1', parent=styles['Normal'],
        fontSize=16, textColor=BLUE, fontName='Helvetica-Bold',
        spaceBefore=8, spaceAfter=4)
    H2 = ParagraphStyle('H2', parent=styles['Normal'],
        fontSize=10, textColor=LBLUE, fontName='Helvetica-Bold',
        spaceBefore=10, spaceAfter=3)
    BODY = ParagraphStyle('BODY', parent=styles['Normal'],
        fontSize=8, textColor=rl_colors.HexColor('#2d3748'),
        spaceAfter=2, leading=11)
    SMALL = ParagraphStyle('SMALL', parent=styles['Normal'],
        fontSize=7, textColor=GRAY, spaceAfter=1)
    CENTER = ParagraphStyle('CENTER', parent=styles['Normal'],
        fontSize=8, alignment=TA_CENTER)

    story = []

    # ── HEADER ────────────────────────────────────────────────────────────────
    story.append(Paragraph("📊 Global TA Dashboard — Morning Bulletin", H1))
    story.append(Paragraph(
        f"{date_str}  •  Timeframe: {freq}  •  Not financial advice",
        SMALL))
    story.append(HRFlowable(width='100%', thickness=1.5,
                             color=LBLUE, spaceAfter=6))

    # Signal summary counts
    counts = df_all["Signal"].value_counts() if "Signal" in df_all.columns else {}
    total  = len(df_all)
    summary_data = [["Signal","Count","%"]]
    for sig_name in ["Strong Buy","Buy","Hold","Sell","Strong Sell"]:
        n = int(counts.get(sig_name, 0))
        pct = f"{round(n/total*100)}%" if total else "0%"
        summary_data.append([sig_name, str(n), pct])

    sum_tbl = Table(summary_data,
        colWidths=[4.5*cm, 2*cm, 2*cm])
    sum_tbl.setStyle(TableStyle([
        ('BACKGROUND',  (0,0),(-1,0), BLUE),
        ('TEXTCOLOR',   (0,0),(-1,0), rl_colors.white),
        ('FONTNAME',    (0,0),(-1,0), 'Helvetica-Bold'),
        ('FONTSIZE',    (0,0),(-1,-1), 7.5),
        ('ROWBACKGROUNDS',(0,1),(-1,-1),[ACCENT, rl_colors.white]),
        ('GRID',        (0,0),(-1,-1), 0.4, rl_colors.HexColor('#AABBD0')),
        ('PADDING',     (0,0),(-1,-1), 3),
        ('ALIGN',       (1,0),(-1,-1), 'CENTER'),
    ]))
    story.append(sum_tbl)
    story.append(Spacer(1, 0.3*cm))

    def section_table(title, rows, col_widths, color=BLUE):
        if not rows: return
        story.append(Paragraph(title, H2))
        hdr = ["Ticker","Name","Signal","Score","ATR%","RSI Div","MACD Div"]
        data = [hdr] + rows
        tbl = Table(data, colWidths=col_widths)
        tbl.setStyle(TableStyle([
            ('BACKGROUND',     (0,0),(-1,0), color),
            ('TEXTCOLOR',      (0,0),(-1,0), rl_colors.white),
            ('FONTNAME',       (0,0),(-1,0), 'Helvetica-Bold'),
            ('FONTSIZE',       (0,0),(-1,-1), 7),
            ('ROWBACKGROUNDS', (0,1),(-1,-1),[ACCENT, rl_colors.white]),
            ('GRID',           (0,0),(-1,-1), 0.4, rl_colors.HexColor('#AABBD0')),
            ('PADDING',        (0,0),(-1,-1), 3),
            ('ALIGN',          (3,0),(-1,-1), 'CENTER'),
        ]))
        story.append(KeepTogether([tbl, Spacer(1, 0.2*cm)]))

    def make_rows(df_sub):
        rows = []
        for _, r in df_sub.iterrows():
            atr_pct = f"{r['ATR%']:.2f}%" if pd.notna(r.get('ATR%')) else "—"
            rows.append([
                str(r.get("Ticker",""))[:12],
                str(r.get("Name",""))[:20],
                str(r.get("Signal","")),
                str(r.get("Score","")),
                atr_pct,
                _div_str_plain(r.get("RSI Div")),
                _div_str_plain(r.get("MACD Div")),
            ])
        return rows

    cw = [2.5*cm, 4.2*cm, 2.3*cm, 1.3*cm, 1.5*cm, 2.5*cm, 2.5*cm]

    # ── SIGNAL CHANGES ────────────────────────────────────────────────────────
    if changes:
        story.append(Paragraph("🔄 Signal Changes Since Last Run", H2))
        chg_data = [["Ticker","Name","Previous","Current","Δ Score"]]
        for ch in changes[:10]:
            delta = ch['curr_score'] - ch['prev_score']
            chg_data.append([
                ch['ticker'][:12],
                ch.get('name','')[:20],
                ch['prev_sig'],
                ch['curr_sig'],
                f"{delta:+d}",
            ])
        chg_tbl = Table(chg_data, colWidths=[2.5*cm,4*cm,2.5*cm,2.5*cm,1.8*cm])
        chg_tbl.setStyle(TableStyle([
            ('BACKGROUND',     (0,0),(-1,0), LBLUE),
            ('TEXTCOLOR',      (0,0),(-1,0), rl_colors.white),
            ('FONTNAME',       (0,0),(-1,0), 'Helvetica-Bold'),
            ('FONTSIZE',       (0,0),(-1,-1), 7),
            ('ROWBACKGROUNDS', (0,1),(-1,-1),[ACCENT, rl_colors.white]),
            ('GRID',           (0,0),(-1,-1), 0.4, rl_colors.HexColor('#AABBD0')),
            ('PADDING',        (0,0),(-1,-1), 3),
        ]))
        story.append(KeepTogether([chg_tbl, Spacer(1, 0.2*cm)]))

    # ── STRONG BUY ────────────────────────────────────────────────────────────
    if "Signal" in df_all.columns:
        sb_df = df_all[df_all["Signal"] == "Strong Buy"].sort_values("Score", ascending=False).head(10)
        if not sb_df.empty:
            section_table("🟢 Strong Buy Candidates", make_rows(sb_df), cw, GREEN)

        # ── STRONG SELL ───────────────────────────────────────────────────────
        ss_df = df_all[df_all["Signal"] == "Strong Sell"].sort_values("Score").head(10)
        if not ss_df.empty:
            section_table("🔴 Strong Sell Candidates", make_rows(ss_df), cw, RED)

    # ── ACTIVE DIVERGENCES ────────────────────────────────────────────────────
    div_rows = []
    if "RSI Div" in df_all.columns and "MACD Div" in df_all.columns:
        for _, r in df_all.iterrows():
            rd = r.get("RSI Div") or {}
            md = r.get("MACD Div") or {}
            rt = rd.get("type") if isinstance(rd, dict) else None
            mt = md.get("type") if isinstance(md, dict) else None
            if rt or mt:
                atr_pct = f"{r['ATR%']:.2f}%" if pd.notna(r.get('ATR%')) else "—"
                div_rows.append([
                    str(r.get("Ticker",""))[:12],
                    str(r.get("Name",""))[:20],
                    str(r.get("Signal","")),
                    str(r.get("Score","")),
                    atr_pct,
                    _div_str_plain(rd),
                    _div_str_plain(md),
                ])
    if div_rows:
        section_table(
            "📐 Active Divergences (RSI and/or MACD)",
            div_rows[:15], cw, AMBER)

    # ── FOOTER ────────────────────────────────────────────────────────────────
    story.append(Spacer(1, 0.4*cm))
    story.append(HRFlowable(width='100%', thickness=0.5,
                             color=GRAY, spaceAfter=4))
    story.append(Paragraph(
        f"Generated: {datetime.now().strftime('%d %b %Y %H:%M')}  •  "
        "Data: Yahoo Finance  •  For research purposes only",
        SMALL))

    doc.build(story)
    return buf.getvalue()

# ─────────────────────────────────────────────────────────────────────────────
# EXCEL EXPORT
# ─────────────────────────────────────────────────────────────────────────────

def build_excel(df_all: pd.DataFrame) -> bytes:
    """Build a formatted Excel workbook from the screener DataFrame."""
    buf = io.BytesIO()
    # Convert div dicts to strings first
    exp = df_all.copy()
    for dc in ["RSI Div","MACD Div"]:
        if dc in exp.columns:
            exp[dc] = exp[dc].apply(_div_str_plain)
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "TA Screener"

        # Header style
        hdr_fill = PatternFill("solid", fgColor="1F3864")
        hdr_font = Font(color="FFFFFF", bold=True, size=9)
        thin     = Side(style='thin', color='AABBD0')
        border   = Border(left=thin, right=thin, top=thin, bottom=thin)

        cols = [c for c in exp.columns]
        ws.append(cols)
        for cell in ws[1]:
            cell.fill = hdr_fill; cell.font = hdr_font
            cell.alignment = Alignment(horizontal='center')
            cell.border = border

        # Signal fill colors
        sig_fills = {
            "Strong Buy":  PatternFill("solid", fgColor="003d1f"),
            "Buy":         PatternFill("solid", fgColor="0d2e2a"),
            "Hold":        PatternFill("solid", fgColor="3a3000"),
            "Sell":        PatternFill("solid", fgColor="3a1a00"),
            "Strong Sell": PatternFill("solid", fgColor="3a0000"),
        }
        sig_fonts = {
            "Strong Buy":  Font(color="00C853", size=8),
            "Buy":         Font(color="26A69A", size=8),
            "Hold":        Font(color="FFD740", size=8),
            "Sell":        Font(color="FF6D00", size=8),
            "Strong Sell": Font(color="FF5252", size=8),
        }

        for row_data in exp.itertuples(index=False):
            ws.append(list(row_data))
            row_idx = ws.max_row
            sig_val = str(getattr(row_data, "Signal", ""))
            for ci, cell in enumerate(ws[row_idx]):
                cell.border = border
                cell.alignment = Alignment(horizontal='center')
                cell.font = Font(size=8)
                # Color the Signal column
                if cols[ci] == "Signal" and sig_val in sig_fills:
                    cell.fill = sig_fills[sig_val]
                    cell.font = sig_fonts[sig_val]
                # Color Score
                if cols[ci] == "Score":
                    try:
                        sv = float(cell.value)
                        if sv >= 2:   cell.font = Font(color="26A69A", bold=True, size=8)
                        elif sv <= -2: cell.font = Font(color="EF5350", bold=True, size=8)
                        else:         cell.font = Font(color="FFD740", size=8)
                    except: pass
                # Color Change%
                if cols[ci] == "Change%":
                    try:
                        sv = float(cell.value)
                        cell.font = Font(color="26A69A" if sv>=0 else "EF5350", size=8)
                    except: pass

        # Auto-width columns
        for ci, col_name in enumerate(cols, 1):
            max_len = max(len(str(col_name)),
                          max((len(str(v)) for v in exp[col_name]), default=0))
            ws.column_dimensions[get_column_letter(ci)].width = min(max_len + 2, 30)

        # Freeze header row
        ws.freeze_panes = "A2"

        wb.save(buf)
    except ImportError:
        # Fallback: plain CSV in the buffer (openpyxl not installed)
        buf.write(exp.to_csv(index=False).encode())
    return buf.getvalue()

# ─────────────────────────────────────────────────────────────────────────────
# SIDEBAR
# ─────────────────────────────────────────────────────────────────────────────

def sidebar() -> dict:
    with st.sidebar:
        st.markdown("## ⚙️ Settings")

        st.markdown('<div class="ss">Timeframe</div>', unsafe_allow_html=True)
        freq = st.selectbox("tf", list(FREQ_OPTIONS.keys()),
                            index=1, label_visibility="collapsed")

        st.markdown('<div class="ss">Indicator Parameters</div>',
                    unsafe_allow_html=True)
        with st.expander("Customize", expanded=False):
            rsi_len   = st.slider("RSI period",      7, 30, 14)
            macd_fast = st.slider("MACD fast",        5, 20, 12)
            macd_slow = st.slider("MACD slow",       15, 50, 26)
            macd_sig  = st.slider("MACD signal",      3, 15,  9)
            cci_len   = st.slider("CCI period",      10, 50, 20)
            mom_len   = st.slider("Momentum period",  5, 30, 10)
            stoch_k   = st.slider("Stochastic %K",   5, 30, 14)
            stoch_d   = st.slider("Stochastic %D",   2, 10,  3)
            willr_len = st.slider("Williams %R",      5, 30, 14)

        st.markdown('<div class="ss">Page 1 — Market Selection</div>',
                    unsafe_allow_html=True)

        def init_group(key, items, default_checked):
            if f"exp_{key}" not in st.session_state:
                st.session_state[f"exp_{key}"] = False
            for name in items:
                if f"item_{key}_{name}" not in st.session_state:
                    st.session_state[f"item_{key}_{name}"] = default_checked

        def asset_group(label, key, items, default_checked=True):
            init_group(key, items, default_checked)
            expanded = st.session_state[f"exp_{key}"]
            c_arr, c_chk, c_lbl = st.columns([0.12, 0.12, 0.76])
            with c_arr:
                arrow = "▼" if expanded else "▶"
                if st.button(arrow, key=f"_arr_{key}"):
                    st.session_state[f"exp_{key}"] = not expanded
            def on_master_change(k=key, it=items):
                new_val = st.session_state[f"_hdr_{k}"]
                for n in it:
                    st.session_state[f"item_{k}_{n}"] = new_val
            any_on = any(st.session_state[f"item_{key}_{n}"] for n in items)
            with c_chk:
                st.checkbox("", value=any_on, key=f"_hdr_{key}",
                            on_change=on_master_change,
                            label_visibility="collapsed")
            with c_lbl:
                st.markdown(f"**{label}**")
            selected = []
            if st.session_state[f"exp_{key}"]:
                _, col = st.columns([0.12, 0.88])
                with col:
                    for name in items:
                        checked = st.checkbox(name,
                            value=st.session_state[f"item_{key}_{name}"],
                            key=f"_chk_{key}_{name}")
                        st.session_state[f"item_{key}_{name}"] = checked
                        if checked: selected.append(name)
            else:
                for name in items:
                    if st.session_state[f"item_{key}_{name}"]:
                        selected.append(name)
            return selected

        sel_idx = asset_group("Indices",     "idx", list(INDICES.keys()),     True)
        sel_com = asset_group("Commodities", "com", list(COMMODITIES.keys()), True)
        sel_fx  = asset_group("Forex",       "fx",  list(FOREX.keys()),       True)
        sel_cr  = asset_group("Crypto",      "cr",  list(CRYPTO.keys()),      True)

        st.markdown("---")
        if st.button("🗑️ Clear cache", use_container_width=True):
            st.cache_data.clear()
            for k in ["screener_p1","screener_p2"]:
                st.session_state.pop(k, None)
            st.rerun()
        st.caption(f"Cache: 1h  •  {datetime.now().strftime('%H:%M %d %b')}")

    p = dict(
        freq=freq,
        period=FREQ_OPTIONS[freq]["period"],
        interval=FREQ_OPTIONS[freq]["interval"],
        rsi_len=rsi_len, macd_fast=macd_fast, macd_slow=macd_slow,
        macd_sig=macd_sig, cci_len=cci_len, mom_len=mom_len,
        stoch_k=stoch_k, stoch_d=stoch_d, willr_len=willr_len,
        sel_idx=sel_idx, sel_com=sel_com, sel_fx=sel_fx, sel_cr=sel_cr,
    )
    return p

# ─────────────────────────────────────────────────────────────────────────────
# SHARED TABLE RENDERER
# ─────────────────────────────────────────────────────────────────────────────

def render_table(df_all: pd.DataFrame, table_key: str = "tbl"):
    f1,f2,f3,f4 = st.columns([2,2,2,2])
    with f1: tf = st.multiselect("Type",   sorted(df_all["Type"].unique()), default=[], key=f"tf_{table_key}")
    with f2: sf = st.multiselect("Signal", ["Strong Buy","Buy","Hold","Sell","Strong Sell"], default=[], key=f"sf_{table_key}")
    with f3: sb = st.selectbox("Sort by",  ["Score","Change%","Ticker","Type","ATR%"], key=f"sb_{table_key}")
    with f4: asc = st.radio("Order", ["Desc","Asc"], horizontal=True, key=f"asc_{table_key}") == "Asc"

    disp = df_all.copy()
    if tf: disp = disp[disp["Type"].isin(tf)]
    if sf: disp = disp[disp["Signal"].isin(sf)]
    if sb in disp.columns:
        disp = disp.sort_values(sb, ascending=asc).reset_index(drop=True)

    counts = df_all["Signal"].value_counts()
    bc = st.columns(5)
    for i, sig in enumerate(["Strong Buy","Buy","Hold","Sell","Strong Sell"]):
        cnt = counts.get(sig,0); col = SIG_COLORS[sig]
        pct = round(cnt/len(df_all)*100) if len(df_all) else 0
        with bc[i]:
            st.markdown(
                f'<div class="mc"><div class="lb">{sig}</div>'
                f'<div class="vl" style="color:{col}">{cnt}</div>'
                f'<div class="sb">{pct}%</div></div>',
                unsafe_allow_html=True)
    st.markdown("<br>", unsafe_allow_html=True)

    ind_cols = ["RSI","MACD","CCI","Momentum%","Stochastic","Williams%R"]
    disp_show = disp.copy()
    for dc in ["RSI Div","MACD Div"]:
        if dc in disp_show.columns:
            disp_show[dc] = disp_show[dc].apply(_div_str_plain)

    show = ["Ticker","Name","Type","Region","Price","Change%","Signal","Score",
            "ATR","ATR%","RSI Div","MACD Div"] + ind_cols
    show = [c for c in show if c in disp_show.columns]

    def cs(v): c=SIG_COLORS.get(v,""); return f"color:{c};font-weight:600;" if c else ""
    def cc(v):
        if pd.isna(v): return ""
        return "color:#26A69A;" if v>=0 else "color:#EF5350;"
    def csc(v):
        if pd.isna(v): return ""
        if v>=2:  return "color:#26A69A;font-weight:600;"
        if v<=-2: return "color:#EF5350;font-weight:600;"
        return "color:#FFD740;"
    def cdiv(v):
        if not isinstance(v,str): return ""
        if "Bull" in v: return "color:#00C853;"
        if "Bear" in v: return "color:#D50000;"
        return "color:#555555;"

    div_cols = [c for c in ["RSI Div","MACD Div"] if c in disp_show.columns]
    styled = disp_show[show].style.map(cs,subset=["Signal"]).map(cc,subset=["Change%"]).map(csc,subset=["Score"])
    if div_cols: styled = styled.map(cdiv, subset=div_cols)
    styled = styled.format({"Price":"{:.2f}","Change%":"{:+.2f}%",
                             "ATR":"{:.4f}","ATR%":"{:.2f}%"}, na_rep="—")
    st.dataframe(styled, use_container_width=True, height=530)

    # ── Export buttons ────────────────────────────────────────────────────────
    ex1, ex2 = st.columns(2)
    csv = disp_show[show].to_csv(index=False)
    with ex1:
        st.download_button("⬇ Export CSV",
            data=csv,
            file_name=f"ta_{datetime.now().strftime('%Y%m%d_%H%M')}.csv",
            mime="text/csv", key=f"csv_{table_key}")
    with ex2:
        excel_bytes = build_excel(disp[show] if all(c in disp.columns for c in show) else disp)
        st.download_button("⬇ Export Excel",
            data=excel_bytes,
            file_name=f"ta_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key=f"xls_{table_key}")

# ─────────────────────────────────────────────────────────────────────────────
# PAGE 1 — MARKET OVERVIEW
# ─────────────────────────────────────────────────────────────────────────────

def page1(p):
    st.markdown("## 📋 Market Overview")
    assets = []
    for name in p["sel_idx"]:
        m = INDICES[name]; assets.append((m["ticker"],name,"Index",m["region"]))
    for name in p["sel_com"]:
        m = COMMODITIES[name]; assets.append((m["ticker"],name,"Commodity",m["region"]))
    for name in p["sel_fx"]:
        m = FOREX[name]; assets.append((m["ticker"],name,"Forex",m["region"]))
    for name in p["sel_cr"]:
        m = CRYPTO[name]; assets.append((m["ticker"],name,"Crypto",m["region"]))

    n = len(assets)
    c1, c2 = st.columns([1,5])
    with c1: run = st.button("▶ Run", type="primary", use_container_width=True)
    with c2: st.caption(f"**{n} assets** · Timeframe: **{p['freq']}** · Score −12→+12 · ATR = 14-day daily")

    if run:
        st.session_state.pop("screener_p1", None)
        rows=[]; bar=st.progress(0); status=st.empty()
        for i,(t,nm,at,reg) in enumerate(assets):
            status.caption(f"⏳ {t} — {nm}  ({i+1}/{n})")
            rows.append(one_row(t,nm,at,reg,p["period"],p["interval"],p))
            bar.progress((i+1)/n)
        bar.empty(); status.empty()
        st.session_state["screener_p1"] = pd.DataFrame(rows)
        # Save to GitHub history
        today_str = datetime.now().strftime("%Y-%m-%d")
        history   = load_signal_history()
        today_sigs = {}
        for row in rows:
            if row.get("Signal","N/A") != "N/A":
                today_sigs[row["Ticker"]] = {
                    "signal": row["Signal"], "score": row["Score"],
                    "name": row["Name"], "atr_pct": row.get("ATR%")
                }
        history[today_str] = today_sigs
        # Keep only last 30 days
        cutoff = (datetime.now() - timedelta(days=30)).strftime("%Y-%m-%d")
        history = {k:v for k,v in history.items() if k >= cutoff}
        save_signal_history(history)

    df_all = st.session_state.get("screener_p1")
    if df_all is None or df_all.empty:
        st.caption("Press **Run** to load data.")
        return
    render_table(df_all, table_key="p1")

# ─────────────────────────────────────────────────────────────────────────────
# PAGE 2 — STOCK SCREENER
# ─────────────────────────────────────────────────────────────────────────────

def page2(p):
    st.markdown("## 📈 Stock Screener")
    c1,c2,c3 = st.columns([1,1,3])
    do_bist = c1.checkbox("🇹🇷 BIST 100", value=True)
    do_sp   = c2.checkbox("🇺🇸 S&P 500",  value=False)
    c3.markdown(
        '<div class="srcnote">📡 <b>BIST 100</b>: Borsa Istanbul CSV (live) · '
        '<b>S&P 500</b>: GitHub/datasets (live) · Fallback if unavailable</div>',
        unsafe_allow_html=True)
    if not do_bist and not do_sp:
        st.info("Select at least one market."); return

    with st.spinner("Checking constituent lists…"):
        bist_u = fetch_bist100_live() if do_bist else {}
        sp_u   = fetch_sp500_live()   if do_sp   else {}

    info = []
    if do_bist:
        src = "live" if bist_u is not BIST100_FALLBACK else "fallback"
        info.append(f"BIST 100: **{len(bist_u)} stocks** ({src})")
    if do_sp:
        src = "live" if sp_u is not SP500_FALLBACK else "fallback"
        info.append(f"S&P 500: **{len(sp_u)} stocks** ({src})")
    st.markdown("  •  ".join(info))

    assets = []
    for ft,nm in bist_u.items(): assets.append((ft,nm,"BIST 100","🇹🇷 Turkey"))
    for tk,nm in sp_u.items():   assets.append((tk,nm,"S&P 500","🇺🇸 US"))

    n = len(assets); est = round(n*0.9/60,1)
    c1,c2 = st.columns([1,5])
    with c1: run2 = st.button("▶ Run Screener", type="primary", use_container_width=True)
    with c2: st.caption(f"**{n} stocks** · Est. ~{est} min · Timeframe: **{p['freq']}** · ATR = 14-day daily")

    if run2:
        st.session_state.pop("screener_p2", None)
        rows=[]; bar=st.progress(0); status=st.empty()
        for i,(ft,nm,at,reg) in enumerate(assets):
            status.caption(f"⏳ {ft} — {nm}  ({i+1}/{n})")
            rows.append(one_row(ft,nm,at,reg,p["period"],p["interval"],p))
            bar.progress((i+1)/n)
        bar.empty(); status.empty()
        st.session_state["screener_p2"] = pd.DataFrame(rows)
        # Save to GitHub history (stocks)
        today_str = datetime.now().strftime("%Y-%m-%d")
        history   = load_signal_history()
        today_sigs = history.get(today_str, {})
        for row in rows:
            if row.get("Signal","N/A") != "N/A":
                today_sigs[row["Ticker"]] = {
                    "signal": row["Signal"], "score": row["Score"],
                    "name": row["Name"], "atr_pct": row.get("ATR%")
                }
        history[today_str] = today_sigs
        cutoff = (datetime.now() - timedelta(days=30)).strftime("%Y-%m-%d")
        history = {k:v for k,v in history.items() if k >= cutoff}
        save_signal_history(history)

    df_all = st.session_state.get("screener_p2")
    if df_all is None or df_all.empty:
        st.caption("Press **Run Screener** to load data."); return
    render_table(df_all, table_key="p2")

# ─────────────────────────────────────────────────────────────────────────────
# PAGE 3 — ASSET DETAIL
# ─────────────────────────────────────────────────────────────────────────────

def page3(p):
    st.markdown("## 🔍 Asset Detail")
    macro = []
    for nm,m in INDICES.items():    macro.append((m["ticker"],nm,"Index",m["region"]))
    for nm,m in COMMODITIES.items():macro.append((m["ticker"],nm,"Commodity",m["region"]))
    for nm,m in FOREX.items():      macro.append((m["ticker"],nm,"Forex",m["region"]))
    for nm,m in CRYPTO.items():     macro.append((m["ticker"],nm,"Crypto",m["region"]))

    stock_assets = []
    if "screener_p2" in st.session_state:
        for _,r in st.session_state["screener_p2"].iterrows():
            stock_assets.append((r["Ticker"],r["Name"],r["Type"],r["Region"]))

    all_assets = macro + stock_assets
    seen,unique = set(),[]
    for item in all_assets:
        if item[0] not in seen: seen.add(item[0]); unique.append(item)

    cats = ["🌍 Indices","🏅 Commodities","💱 Forex","₿ Crypto","📈 Stocks"]
    tabs = st.tabs(cats)

    def assets_in_cat(cat):
        m = {"🌍 Indices":"Index","🏅 Commodities":"Commodity",
             "💱 Forex":"Forex","₿ Crypto":"Crypto"}
        if cat in m:
            return [a for a in unique if a[2]==m[cat]]
        return [a for a in unique if a[2] not in ("Index","Commodity","Forex","Crypto")]

    for tab,cat in zip(tabs,cats):
        with tab:
            pool = assets_in_cat(cat)
            if not pool:
                st.info("Run the Stock Screener (Page 2) first to browse stocks here."
                        if cat=="📈 Stocks" else "No assets in this category.")
                continue
            search = st.text_input("🔍 Search",placeholder="Type name or ticker…",
                key=f"srch_{cat}",label_visibility="collapsed")
            if search:
                q = search.lower()
                pool = [a for a in pool if q in a[0].lower() or q in a[1].lower()]
            if pool:
                opts   = [a[0] for a in pool]
                labels = {a[0]:f"{a[0]}  —  {a[1]}" for a in pool}
                chosen = st.selectbox("asset",opts,format_func=lambda t:labels.get(t,t),
                    key=f"sel_{cat}",label_visibility="collapsed")
                if st.button("Load chart",key=f"load_{cat}",type="primary"):
                    st.session_state["detail_ticker"] = chosen
                    st.session_state["detail_meta"]   = next(a for a in pool if a[0]==chosen)

    if "detail_ticker" not in st.session_state:
        st.info("↑ Select an asset above and press Load chart.")
        return

    full_ticker = st.session_state["detail_ticker"]
    _,name,atype,region = st.session_state["detail_meta"]

    ci,cr = st.columns([5,1])
    with ci: st.markdown(f"**{full_ticker}** — {name} &nbsp;|&nbsp; {atype} &nbsp;|&nbsp; {region}")
    with cr:
        if st.button("🔄 Refresh",use_container_width=True):
            fetch_data.clear(); fetch_atr_daily.clear(); st.rerun()

    with st.spinner(f"Loading {full_ticker}…"):
        df       = fetch_data(full_ticker, p["period"], p["interval"])
        atr_data = fetch_atr_daily(full_ticker)

    if df.empty:
        st.error(f"No data for **{full_ticker}**. Try Refresh."); return

    sig     = compute_signal(df, p, full_series=True, interval=p["interval"])
    price   = float(df["close"].iloc[-1])
    chg_pct = float((df["close"].iloc[-1]/df["close"].iloc[-2]-1)*100) if len(df)>1 else 0.
    chg_col = "#26A69A" if chg_pct>=0 else "#EF5350"
    sc_col  = SIG_COLORS.get(sig["signal"],"#FFD740")
    p_fmt   = f"{price:,.4f}" if price<10 else f"{price:,.2f}"
    atr_str = f"{atr_data['atr']:.4f} ({atr_data['atr_pct']:.2f}%)" if atr_data else "N/A"

    # Header cards (6 now: + ATR)
    c1,c2,c3,c4,c5,c6 = st.columns(6)
    for col,(lbl,val,clr) in zip(
        [c1,c2,c3,c4,c5,c6],
        [("Asset",name,"#FAFAFA"),("Type",atype,"#63b3ed"),
         ("Price",p_fmt,"#FAFAFA"),
         (f"Change ({p['freq']})",f"{chg_pct:+.2f}%",chg_col),
         ("ATR (14D Daily)",atr_str,"#FFD740"),
         ("TA Signal",sig["signal"],sc_col)],
    ):
        with col:
            st.markdown(
                f'<div class="mc"><div class="lb">{lbl}</div>'
                f'<div class="vl" style="color:{clr};font-size:.9rem;">{val}</div></div>',
                unsafe_allow_html=True)

    st.markdown("<br>", unsafe_allow_html=True)

    with st.expander("📊 Score Breakdown", expanded=True):
        SL = {2:"Strong Buy",1:"Buy",0:"Hold",-1:"Sell",-2:"Strong Sell"}
        bc2 = st.columns(6)
        for i,(k,(sc,vl)) in enumerate(sig.get("details",{}).items()):
            cc = SIG_COLORS.get(SL.get(sc,"Hold"),"#FFD740")
            with bc2[i]:
                st.markdown(
                    f'<div class="mc"><div class="lb">{k}</div>'
                    f'<div class="vl" style="color:{cc}">{sc:+d}</div>'
                    f'<div class="sb">{vl}</div></div>',
                    unsafe_allow_html=True)
        st.markdown("<br>", unsafe_allow_html=True)
        total = sig["score"]
        css_cls = sig["signal"].lower().replace(" ","-")
        st.markdown(
            f"**Composite: {total:+d}** / ±12 &nbsp;"
            f'<span class="sig sig-{css_cls}">{sig["signal"]}</span>',
            unsafe_allow_html=True)
        st.progress(int((total+12)/24*100))

        divs = sig.get("divergences",{})
        if divs:
            st.markdown("<br>**Divergences** (RSI & MACD histogram)",
                        unsafe_allow_html=True)
            dc1,dc2 = st.columns(2)
            for col,(key,label) in zip([dc1,dc2],[("rsi","RSI"),("macd","MACD Hist")]):
                d = divs.get(key,{"type":None,"strength":0})
                dtype = d.get("type"); strength = d.get("strength",0)
                disp_lbl,color = DIV_LABELS.get(dtype,DIV_LABELS[None])
                dots = STRENGTH_LABEL.get(strength,"")
                with col:
                    st.markdown(
                        f'<div class="mc"><div class="lb">{label}</div>'
                        f'<div class="vl" style="color:{color};font-size:.9rem;">'
                        f'{disp_lbl} {dots}</div></div>',
                        unsafe_allow_html=True)

    st.markdown("<br>", unsafe_allow_html=True)
    ind  = sig.get("ind_series",{})
    divs = sig.get("divergences",{})
    if ind:
        st.plotly_chart(
            build_charts(df, ind, full_ticker, name, p["interval"],
                         divs=divs, atr_data=atr_data),
            use_container_width=True, config={"displayModeBar":True})
    else:
        st.warning("Not enough data for charts (need ≥ 35 bars).")

    with st.expander("📄 Raw data (last 50 bars)"):
        fmt = "%Y-%m-%d %H:%M" if p["interval"]=="1h" else "%Y-%m-%d"
        sd = df.tail(50).copy(); sd.index = sd.index.strftime(fmt)
        st.dataframe(sd.style.format("{:.4f}"), use_container_width=True)

# ─────────────────────────────────────────────────────────────────────────────
# PAGE 4 — MORNING BULLETIN
# ─────────────────────────────────────────────────────────────────────────────

def page4(p):
    st.markdown("## 📰 Morning Bulletin")
    st.caption(
        "Auto-generated from the most recent screener run. "
        "Run Page 1 and/or Page 2 first to populate this bulletin.")

    # Collect data from both screeners
    frames = []
    if "screener_p1" in st.session_state:
        frames.append(st.session_state["screener_p1"])
    if "screener_p2" in st.session_state:
        frames.append(st.session_state["screener_p2"])

    if not frames:
        st.info("No screener results yet. Go to Page 1 or Page 2 and press Run.")
        return

    df_all = pd.concat(frames, ignore_index=True).drop_duplicates(subset=["Ticker"])

    # Load signal history for change detection
    history      = load_signal_history()
    today_str    = datetime.now().strftime("%Y-%m-%d")
    today_sigs   = {}
    for _,r in df_all.iterrows():
        if r.get("Signal","N/A") != "N/A":
            today_sigs[r["Ticker"]] = {
                "signal":r["Signal"], "score":r["Score"],
                "name":r.get("Name",""), "atr_pct":r.get("ATR%")
            }
    changes = get_signal_changes(history, today_sigs)

    date_str = datetime.now().strftime("%d %B %Y")
    st.markdown(f"### 📅 {date_str}  ·  Timeframe: {p['freq']}")
    st.markdown("---")

    # ── Signal change summary ─────────────────────────────────────────────────
    if changes:
        st.markdown("#### 🔄 Signal Changes Since Last Run")
        cols_h = ["Ticker","Name","Previous","→","Current","Δ Score"]
        chg_rows = []
        for ch in changes[:12]:
            delta = ch["curr_score"] - ch["prev_score"]
            arrow = "⬆" if ch["direction"]=="up" else "⬇"
            clr   = "chg-up" if ch["direction"]=="up" else "chg-dn"
            chg_rows.append(
                f'<tr>'
                f'<td><b>{ch["ticker"]}</b></td>'
                f'<td>{ch.get("name","")[:22]}</td>'
                f'<td style="color:#718096">{ch["prev_sig"]}</td>'
                f'<td>→</td>'
                f'<td><b>{ch["curr_sig"]}</b></td>'
                f'<td class="{clr}">{delta:+d}</td>'
                f'</tr>'
            )
        html = (
            '<table style="width:100%;border-collapse:collapse;font-size:.82rem;">'
            '<thead><tr style="border-bottom:2px solid #2d3748;color:#718096;">'
            + "".join(f"<th style='padding:.4rem .5rem;text-align:left'>{h}</th>" for h in cols_h)
            + '</tr></thead><tbody>'
            + "".join(chg_rows)
            + '</tbody></table>'
        )
        st.markdown(html, unsafe_allow_html=True)
        st.markdown("<br>", unsafe_allow_html=True)
    else:
        st.markdown("#### 🔄 Signal Changes")
        if not history:
            st.caption("No history yet — signals will be compared after the second run.")
        else:
            st.caption("No signal changes detected since the last run.")

    # ── Market sentiment bar ──────────────────────────────────────────────────
    st.markdown("#### 📊 Market Sentiment")
    counts = df_all["Signal"].value_counts()
    total  = len(df_all)
    sc = st.columns(5)
    for i,sig in enumerate(["Strong Buy","Buy","Hold","Sell","Strong Sell"]):
        cnt = counts.get(sig,0); col = SIG_COLORS[sig]
        pct = round(cnt/total*100) if total else 0
        with sc[i]:
            st.markdown(
                f'<div class="mc"><div class="lb">{sig}</div>'
                f'<div class="vl" style="color:{col}">{cnt}</div>'
                f'<div class="sb">{pct}%</div></div>',
                unsafe_allow_html=True)
    st.markdown("<br>", unsafe_allow_html=True)

    # ── Strong Buy candidates ─────────────────────────────────────────────────
    sb_df = df_all[df_all["Signal"]=="Strong Buy"].sort_values("Score",ascending=False)
    if not sb_df.empty:
        st.markdown(f"#### 🟢 Strong Buy  ({len(sb_df)} assets)")
        _bulletin_table(sb_df, "#00C853")

    # ── Strong Sell candidates ────────────────────────────────────────────────
    ss_df = df_all[df_all["Signal"]=="Strong Sell"].sort_values("Score")
    if not ss_df.empty:
        st.markdown(f"#### 🔴 Strong Sell  ({len(ss_df)} assets)")
        _bulletin_table(ss_df, "#D50000")

    # ── Active divergences ────────────────────────────────────────────────────
    if "RSI Div" in df_all.columns and "MACD Div" in df_all.columns:
        def has_div(row):
            rd = row.get("RSI Div") or {}
            md = row.get("MACD Div") or {}
            rt = rd.get("type") if isinstance(rd,dict) else None
            mt = md.get("type") if isinstance(md,dict) else None
            return bool(rt or mt)
        div_df = df_all[df_all.apply(has_div, axis=1)].copy()
        if not div_df.empty:
            # Prioritise: strong classic divergences with high/low score
            div_df["_classic"] = div_df.apply(lambda r: (
                1 if isinstance(r.get("RSI Div"),dict) and
                     r["RSI Div"].get("type") in ("bullish","bearish") else 0
            ) + (
                1 if isinstance(r.get("MACD Div"),dict) and
                     r["MACD Div"].get("type") in ("bullish","bearish") else 0
            ), axis=1)
            div_df = div_df.sort_values(["_classic","Score"], ascending=[False,True])
            st.markdown(f"#### 📐 Active Divergences  ({len(div_df)} assets)")
            _bulletin_table(div_df, "#FF6D00", show_div=True)

    # ── PDF Export ────────────────────────────────────────────────────────────
    st.markdown("---")
    st.markdown("#### 📄 Export Bulletin")
    if st.button("Generate PDF Bulletin", type="primary"):
        with st.spinner("Building PDF…"):
            pdf_bytes = build_bulletin_pdf(df_all, p["freq"], changes, date_str)
        st.download_button(
            "⬇ Download PDF",
            data=pdf_bytes,
            file_name=f"ta_bulletin_{datetime.now().strftime('%Y%m%d')}.pdf",
            mime="application/pdf",
        )


def _bulletin_table(df_sub: pd.DataFrame, header_color: str, show_div: bool = False):
    """Render a compact HTML table for the bulletin page."""
    cols = ["Ticker","Name","Signal","Score","ATR%","Change%"]
    if show_div:
        cols += ["RSI Div","MACD Div"]

    header = "".join(
        f"<th style='padding:.35rem .5rem;text-align:left;font-size:.75rem;"
        f"color:#FAFAFA;'>{c}</th>" for c in cols)
    rows_html = ""
    for _,r in df_sub.head(15).iterrows():
        chg_c = "#26A69A" if (r.get("Change%") or 0) >= 0 else "#EF5350"
        atr_v = f"{r['ATR%']:.2f}%" if pd.notna(r.get("ATR%")) else "—"
        sig_c = SIG_COLORS.get(r.get("Signal","N/A"),"#FFD740")
        cells = [
            f"<b style='color:#63b3ed'>{r.get('Ticker','')}</b>",
            f"{str(r.get('Name',''))[:22]}",
            f"<span style='color:{sig_c};font-weight:600'>{r.get('Signal','')}</span>",
            f"<b>{r.get('Score',0):+d}</b>",
            atr_v,
            f"<span style='color:{chg_c}'>{r.get('Change%',0):+.2f}%</span>",
        ]
        if show_div:
            for dc in ["RSI Div","MACD Div"]:
                dv = r.get(dc)
                ds = _div_str_plain(dv)
                col = "#00C853" if "Bull" in ds else ("#D50000" if "Bear" in ds else "#555")
                cells.append(f"<span style='color:{col}'>{ds}</span>")
        rows_html += "<tr>" + "".join(
            f"<td style='padding:.3rem .5rem;font-size:.78rem;"
            f"border-bottom:1px solid #1a202c;'>{c}</td>" for c in cells) + "</tr>"

    html = (
        f'<table style="width:100%;border-collapse:collapse;">'
        f'<thead><tr style="background:{header_color}22;border-bottom:2px solid {header_color};">'
        f'{header}</tr></thead><tbody>{rows_html}</tbody></table>'
    )
    st.markdown(html, unsafe_allow_html=True)
    st.markdown("<br>", unsafe_allow_html=True)

# ─────────────────────────────────────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────────────────────────────────────

def main():
    st.markdown(CSS, unsafe_allow_html=True)
    st.markdown(
        '<div class="hdr"><h1>📊 Global Technical Analysis Dashboard</h1>'
        '<p>Indices · Commodities · Forex · Crypto · BIST 100 · S&P 500 · '
        '6 indicators · ATR · Divergences · Morning Bulletin</p></div>',
        unsafe_allow_html=True)

    p  = sidebar()
    pg = st.radio("nav",
        ["📋 Market Overview","📈 Stock Screener",
         "🔍 Asset Detail","📰 Morning Bulletin"],
        horizontal=True, label_visibility="hidden")
    st.markdown("---")

    if   "Overview"  in pg: page1(p)
    elif "Stock"     in pg: page2(p)
    elif "Asset"     in pg: page3(p)
    else:                   page4(p)

    st.markdown(
        f'<div class="footer">Data: Yahoo Finance · Borsa Istanbul · GitHub/datasets  '
        f'•  Indicators: pandas-ta  •  Not financial advice  '
        f'•  {datetime.now().strftime("%d %b %Y %H:%M")}</div>',
        unsafe_allow_html=True)

if __name__ == "__main__":
    main()
